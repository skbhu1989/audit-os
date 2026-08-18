import io
from uuid import UUID
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from ..db import tenant_conn
from ..deps import get_current_user, CurrentUser

router = APIRouter(prefix="/engagements/{engagement_id}/export", tags=["reporting"])

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(60, max_len + 2)
    return ws


@router.get("/audit-file.xlsx")
async def export_audit_file(engagement_id: UUID, user: CurrentUser = Depends(get_current_user)):
    async with tenant_conn(user.firm_id) as conn:
        eng = await conn.fetchrow(
            "select financial_year, reporting_date from engagement where id=$1", engagement_id
        )
        if not eng:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Engagement not found")

        gst_rows = await conn.fetch(
            """select r.recon_type, e.document_no, e.period, e.party_name, e.books_amount, e.return_amount,
                      e.difference, e.reason, e.risk_level
               from reconciliation_exception e join reconciliation_run r on r.id=e.run_id
               where r.engagement_id=$1 and r.recon_type like 'GST_%' order by e.risk_level""",
            engagement_id,
        )
        tds_rows = await conn.fetch(
            """select e.document_no as section, e.books_amount as deducted, e.return_amount as paid,
                      e.reason, e.risk_level
               from reconciliation_exception e join reconciliation_run r on r.id=e.run_id
               where r.engagement_id=$1 and r.recon_type='TDS_RECONCILIATION' order by e.risk_level""",
            engagement_id,
        )
        payroll_rows = await conn.fetch(
            """select r.recon_type, e.period, e.books_amount as liability, e.return_amount as paid,
                      e.reason, e.risk_level
               from reconciliation_exception e join reconciliation_run r on r.id=e.run_id
               where r.engagement_id=$1 and r.recon_type like 'PAYROLL_%' order by e.risk_level""",
            engagement_id,
        )
        journals = await conn.fetch(
            """select id, posted_date, posted_by, amount, risk_level, risk_reasons
               from journal where engagement_id=$1 and risk_level is not null order by risk_score desc""",
            engagement_id,
        )
        calendar_rows = await conn.fetch(
            """select statutory_type, filing_or_payment, period, due_date, actual_date, status
               from compliance_calendar_item where engagement_id=$1 order by due_date""",
            engagement_id,
        )
        caro_rows = await conn.fetch(
            """select c.clause_no, cl.title, c.applicability, c.data_status, c.status,
                      coalesce(c.final_response, c.draft_response, c.data_gap_reason) as response
               from caro_assessment c join caro_clause cl on cl.clause_no=c.clause_no
               where c.engagement_id=$1
               order by array_position(array['i','ii','iii','iv','v','vi','vii','viii','ix','x',
                 'xi','xii','xiii','xiv','xv','xvi','xvii','xviii','xix','xx','xxi'], c.clause_no)""",
            engagement_id,
        )

    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    _write_sheet(wb, "GST Reconciliation", ["Type", "Document No", "Period", "Party", "Books Amt", "Return Amt", "Difference", "Reason", "Risk"],
                 [[r["recon_type"], r["document_no"], r["period"], r["party_name"],
                   float(r["books_amount"]) if r["books_amount"] is not None else None,
                   float(r["return_amount"]) if r["return_amount"] is not None else None,
                   float(r["difference"]) if r["difference"] is not None else None,
                   r["reason"], r["risk_level"]] for r in gst_rows])

    _write_sheet(wb, "TDS Reconciliation", ["Section", "Deducted", "Paid", "Reason", "Risk"],
                 [[r["section"], float(r["deducted"]) if r["deducted"] is not None else None,
                   float(r["paid"]) if r["paid"] is not None else None, r["reason"], r["risk_level"]] for r in tds_rows])

    _write_sheet(wb, "Payroll Reconciliation", ["Scheme", "Period", "Liability", "Paid", "Reason", "Risk"],
                 [[r["recon_type"].split("_")[1], r["period"],
                   float(r["liability"]) if r["liability"] is not None else None,
                   float(r["paid"]) if r["paid"] is not None else None, r["reason"], r["risk_level"]] for r in payroll_rows])

    _write_sheet(wb, "Journal Entry Testing", ["Journal ID", "Date", "Posted By", "Amount", "Risk Level", "Reasons"],
                 [[str(r["id"])[:8], str(r["posted_date"]), r["posted_by"], float(r["amount"]),
                   r["risk_level"], "; ".join(r["risk_reasons"] or [])] for r in journals])

    _write_sheet(wb, "Compliance Calendar", ["Statutory Type", "Filing/Payment", "Period", "Due Date", "Actual Date", "Status"],
                 [[r["statutory_type"], r["filing_or_payment"], r["period"], str(r["due_date"]),
                   str(r["actual_date"]) if r["actual_date"] else "", r["status"]] for r in calendar_rows])

    _write_sheet(wb, "CARO", ["Clause", "Title", "Applicability", "Data Status", "Sign-off Status", "Response (draft or final)"],
                 [[r["clause_no"], r["title"], r["applicability"], r["data_status"], r["status"], r["response"]] for r in caro_rows])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"audit-file-{eng['financial_year']}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
